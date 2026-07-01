"""CalSim 19-basin region registry: names, flow ratios, and dissolved watershed polygons.

Ties together two external sources that never reference each other directly:

- ``data/CalSim-19Basins-FlowContributionPercentage.xlsx`` (``AreaRatio_FlowRatio`` sheet) defines
  the 19 named basin regions (ids 0-18) plus the Delta (id 19), each with a Flow Ratio -- the same
  weighting scheme the existing ``cv-flow-weighted`` LOCA-2 product already uses. Its
  ``RimInflow_in_CalSim3`` sheet separately gives a ``CalSim node -> region id`` crosswalk (row
  index 1 = node pathnames ``I_<Connect_No>``, row index 3 = the region id per node -- verified
  against the geopackage's ``CalSim3_Merged`` Connect_No set, 199/200 match).
- A CalSim3 geopackage (``.gpkg``, a SQLite file) provides the actual watershed polygons: rim
  inflow polygons in ``CalSim3_Merged`` (join via the crosswalk above), and fragmented Sacramento /
  San Joaquin valley-floor sub-areas plus a single named Tulare Lake Basin polygon in
  ``CalSim3_And_GooseLake``.

No Goose Lake (id 0) or Delta (id 19) polygon exists in either geopackage table -- both are simply
left with ``polygon=None`` (Goose Lake has flow ratio 0 so this is inconsequential; Delta is
excluded by design, see ``historical.py``).

Geometry blobs are read via stdlib ``sqlite3`` (no fiona/pyogrio/geopandas) and decoded with
``shapely.wkb`` after stripping the GeoPackage binary header (magic ``GP`` + version + flags byte,
which encodes an optional envelope to skip -- see the OGC GeoPackage spec, clause 2.1.3).
"""

from __future__ import annotations

import re
import sqlite3
from dataclasses import dataclass, replace
from pathlib import Path

import openpyxl
from shapely import wkb as shapely_wkb
from shapely.geometry.base import BaseGeometry
from shapely.ops import unary_union

_ENVELOPE_BYTES = {0: 0, 1: 32, 2: 48, 3: 48, 4: 64}

# Sacramento WBAs are numbered <= 26 (with -North/-South suffixes); San Joaquin WBAs are numbered
# >= 60. The ranges are disjoint, so a single cutoff unambiguously routes each valley-floor
# fragment to ValleyFloor_Sac (region 12) or ValleyFloor_SJR (region 13).
_SAC_SJR_WBA_CUTOFF = 40
VALLEYFLOOR_SAC_ID = 12
VALLEYFLOOR_SJR_ID = 13
TULARE_ID = 14

_WBA_RE = re.compile(r"WBA\s*(\d+)")


@dataclass(frozen=True)
class BasinRegion:
    id: int
    name: str
    flow_ratio: float
    polygon: BaseGeometry | None = None  # None if unresolved (Goose Lake, Delta)


def _parse_gpkg_geometry(blob: bytes):
    """Strip the GeoPackage binary header and decode the remaining WKB via shapely."""
    if blob[0:2] != b"GP":
        raise ValueError(f"not a GeoPackage geometry blob (magic={blob[0:2]!r})")
    flags = blob[3]
    envelope_code = (flags >> 1) & 0x07
    header_len = 8 + _ENVELOPE_BYTES[envelope_code]
    return shapely_wkb.loads(blob[header_len:])


def read_gpkg_table(gpkg_path: Path, table: str, extra_cols: tuple[str, ...] = ()) -> list[dict]:
    """Read a GeoPackage feature table's geometry (as shapely) plus requested attribute columns."""
    con = sqlite3.connect(str(gpkg_path))
    try:
        cur = con.cursor()
        cols = ["geom", *extra_cols]
        cur.execute(f"SELECT {', '.join(cols)} FROM {table}")
        rows = []
        for row in cur.fetchall():
            geom_blob = row[0]
            geom = _parse_gpkg_geometry(geom_blob) if geom_blob else None
            rec = dict(zip(extra_cols, row[1:]))
            rec["geom"] = geom
            rows.append(rec)
        return rows
    finally:
        con.close()


def load_rim_crosswalk(xlsx_path: Path) -> dict[str, int]:
    """``CalSim Connect_No -> basin region id`` for the ~199 rim-watershed inflow nodes."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["RimInflow_in_CalSim3"]
    rows = list(ws.iter_rows(values_only=True))
    part_b, part_d = rows[1], rows[3]
    assert str(part_b[0]).startswith("Part B"), f"unexpected RimInflow_in_CalSim3 layout: {part_b[0]!r}"
    assert str(part_d[0]).startswith("Part D"), f"unexpected RimInflow_in_CalSim3 layout: {part_d[0]!r}"

    crosswalk: dict[str, int] = {}
    for name, region in zip(part_b, part_d):
        if not isinstance(name, str) or not name.startswith("I_"):
            continue
        if region in (None, ""):
            continue
        crosswalk[name[2:]] = int(region)
    return crosswalk


def load_flow_ratios(xlsx_path: Path) -> dict[int, BasinRegion]:
    """The 19 named basin regions (ids 0-18) plus Delta (id 19), with their Flow Ratio weights."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["AreaRatio_FlowRatio"]
    regions: dict[int, BasinRegion] = {}
    for row in ws.iter_rows(values_only=True):
        if len(row) < 14:
            continue
        region_id, name, flow_ratio = row[8], row[9], row[13]
        if not isinstance(region_id, (int, float)) or name is None or flow_ratio is None:
            continue
        region_id = int(region_id)
        if not (0 <= region_id <= 19):
            continue
        regions[region_id] = BasinRegion(id=region_id, name=str(name).strip(), flow_ratio=float(flow_ratio))
    return regions


def _parse_wba_number(remarks: str) -> int | None:
    m = _WBA_RE.search(remarks or "")
    return int(m.group(1)) if m else None


def build_basin_polygons(gpkg_path: Path, xlsx_path: Path) -> dict[int, BaseGeometry]:
    """Dissolved (unioned) polygon per resolvable basin region id.

    Rim regions come from ``CalSim3_Merged`` via the Connect_No crosswalk; the Sacramento/SJR
    valley floor and Tulare Lake Basin come from the ``Valley``-type and named rows of
    ``CalSim3_And_GooseLake``. Goose Lake (id 0) and Delta (id 19) have no polygon in either table
    and are simply absent from the returned dict.
    """
    crosswalk = load_rim_crosswalk(xlsx_path)
    geoms_by_region: dict[int, list] = {}

    merged = read_gpkg_table(gpkg_path, "CalSim3_Merged", extra_cols=("Connect_No",))
    for rec in merged:
        if rec["geom"] is None:
            continue
        region_id = crosswalk.get((rec["Connect_No"] or "").strip())
        if region_id is not None:
            geoms_by_region.setdefault(region_id, []).append(rec["geom"])

    goose = read_gpkg_table(gpkg_path, "CalSim3_And_GooseLake", extra_cols=("Type", "Remarks"))
    for rec in goose:
        if rec["geom"] is None:
            continue
        remarks = rec["Remarks"] or ""
        if rec["Type"] == "Valley":
            wba = _parse_wba_number(remarks)
            if wba is None:
                continue
            region_id = VALLEYFLOOR_SAC_ID if wba < _SAC_SJR_WBA_CUTOFF else VALLEYFLOOR_SJR_ID
            geoms_by_region.setdefault(region_id, []).append(rec["geom"])
        elif "Tulare Lake Basin" in remarks:
            geoms_by_region.setdefault(TULARE_ID, []).append(rec["geom"])

    return {region_id: unary_union(geoms) for region_id, geoms in geoms_by_region.items()}


def load_basin_registry(gpkg_path: Path, xlsx_path: Path) -> dict[int, BasinRegion]:
    """The full basin-region registry: name, flow ratio, and (where resolvable) polygon."""
    flow_ratios = load_flow_ratios(xlsx_path)
    polygons = build_basin_polygons(gpkg_path, xlsx_path)
    return {
        region_id: replace(region, polygon=polygons.get(region_id))
        for region_id, region in flow_ratios.items()
    }
